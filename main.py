from openpyxl import Workbook
from io import BytesIO
import streamlit as st

st.set_page_config(page_title="Civil AI Assistant", page_icon="🏗️", layout="wide")

st.title("🏗️ Civil AI Assistant")
st.subheader("WhatsApp to Excel")

st.write("WhatsApp ka message yahan paste karein.")

message = st.text_area(
    "Paste WhatsApp Message",
    height=300,
    placeholder="WhatsApp message yahan paste karein..."
)
if st.button("Generate Excel"):
    if message.strip() == "":
        st.warning("Pehle WhatsApp message paste karein.")
    else:
        st.success("Message receive ho gaya.")

        st.subheader("AI Output")

        if "Earthwork" in message or "earthwork" in message:
            st.info("Work Type: Earthwork")
        else:
            st.info("Work Type: Structure")

        wb = Workbook()
ws = wb.active
ws.title = "Requests"

ws["A1"] = "WhatsApp Message"
ws["A2"] = message

buffer = BytesIO()
wb.save(buffer)

st.download_button(
    "📥 Download Excel",
    data=buffer.getvalue(),
    file_name="Civil_Request.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
st.divider()

st.subheader("Download Excel")

if st.button("Create Excel File"):
    st.success("Excel file tayyar hai. (Next step mein actual Excel generate hogi)")
